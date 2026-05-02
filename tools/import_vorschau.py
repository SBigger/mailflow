"""
Dry-Run-Vorschau für den Adress-Import.

Liest:
- Desktop/Abgleich_Adressen.xlsx          (vom User bereinigte Quelle)
- %TEMP%/crm_snapshot.json                (vom export-crm-snapshot Edge Function)

Schreibt:
- Desktop/Import_Vorschau.xlsx            (rein lesbar – simuliert was passieren würde)

Wichtig: Schreibt NICHTS in die DB. Reine Vorschau.

Sheets im Output:
  1. Uebersicht
  2. NEU_Privatkunden          (INSERT-Kandidaten aus 'Privatkunden_fehlen')
  3. UPDATE_Kontakte           (contact_persons Append-Kandidaten)
  4. SKIP_Kontakte             (Zeilen die übersprungen werden – kein Match, kein Tel/Mail, ...)
"""

import json
import os
import re
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

# ── Pfade ────────────────────────────────────────────────
DESKTOP   = Path(r"C:\Users\SaschaBigger\Artis Treuhand GmbH\OneDrive - Artis Treuhand GmbH\Desktop")
SRC_XLSX  = DESKTOP / "Abgleich_Adressen.xlsx"
OUT_XLSX  = DESKTOP / "Import_Vorschau.xlsx"
SNAPSHOT  = Path(os.environ.get("TEMP", r"C:\Users\SaschaBigger\AppData\Local\Temp")) / "crm_snapshot.json"

# ── Helpers ──────────────────────────────────────────────
def norm(s):
    if s is None: return ""
    return re.sub(r"\s+", " ", str(s)).strip()

def norm_lower(s):
    return norm(s).lower()

def is_red_cell(cell):
    f = cell.fill
    if not (f and f.fgColor and f.fgColor.rgb): return False
    rgb = str(f.fgColor.rgb)
    try:
        rr = int(rgb[2:4], 16); gg = int(rgb[4:6], 16); bb = int(rgb[6:8], 16)
        return rr > 200 and gg < 180 and bb < 180 and (rr - gg) > 40
    except Exception:
        return False

def row_is_red(row_cells):
    return any(is_red_cell(c) for c in row_cells)

# ── Excel einlesen (bereinigt vom User) ──────────────────
def read_sheet(wb, name):
    if name not in wb.sheetnames:
        return []
    ws = wb[name]
    rows = list(ws.iter_rows())
    if not rows: return []
    headers = [c.value for c in rows[0]]
    out = []
    for row in rows[1:]:
        if all(c.value in (None, "") for c in row):
            continue
        if row_is_red(row):
            continue
        d = {}
        for h, c in zip(headers, row):
            d[h] = c.value
        # Komplett leere Zeilen (nur ID)? Skip wenn gar nichts Inhaltliches
        if not any(norm(v) for k, v in d.items() if k not in ("M-Files-ID", "CRM-ID Firma", "CRM-ID")):
            continue
        out.append(d)
    return out

# ── CRM laden ────────────────────────────────────────────
def load_crm():
    with open(SNAPSHOT, encoding="utf-8") as f:
        data = json.load(f)
    return data.get("customers", [])

# ── Hauptlogik ───────────────────────────────────────────
def main():
    if not SRC_XLSX.exists():
        raise SystemExit(f"FEHLT: {SRC_XLSX}")
    if not SNAPSHOT.exists():
        raise SystemExit(f"FEHLT: {SNAPSHOT} – bitte export-crm-snapshot erneut aufrufen")

    print(f"Lade Excel: {SRC_XLSX}", flush=True)
    wb_src = openpyxl.load_workbook(SRC_XLSX, data_only=True)

    privat_fehlen   = read_sheet(wb_src, "Privatkunden_fehlen")
    kontakte_erg    = read_sheet(wb_src, "Kontakte_ergaenzen")

    print(f"  Privatkunden_fehlen:  {len(privat_fehlen)}", flush=True)
    print(f"  Kontakte_ergaenzen:   {len(kontakte_erg)}", flush=True)

    print(f"Lade CRM Snapshot: {SNAPSHOT}", flush=True)
    customers = load_crm()
    by_id = {c["id"]: c for c in customers}
    print(f"  {len(customers)} Kunden", flush=True)

    # ── NEU_Privatkunden bauen ──────────────────────────
    neu_privat = []
    for r in privat_fehlen:
        vor = norm(r.get("Vorname"))
        nach = norm(r.get("Nachname"))
        if not (vor or nach): continue
        company = (vor + " " + nach).strip()
        neu_privat.append({
            "company_name":   company,         # CRM-Konvention für Privatkunden
            "vorname":        vor,
            "name":           nach,
            "kunde_typ":      "privat",
            "phone":          norm(r.get("Telefon")),
            "phone2":         norm(r.get("Mobil")),
            "email":          norm(r.get("E-Mail")),
            "strasse":        "",
            "plz":            "",
            "ort":            "",
            "geburtstag":     norm(r.get("Geburtstag")),
            "M-Files-ID":     norm(r.get("M-Files-ID")),
            "_Quelle":        f"{r.get('Adresse') or ''}",
            "_Aktion":        "INSERT customers",
        })

    # ── UPDATE_Kontakte bauen ───────────────────────────
    update_kontakte = []
    skip_kontakte   = []
    for r in kontakte_erg:
        crm_id = norm(r.get("CRM-ID Firma"))
        vor    = norm(r.get("Vorname"))
        nach   = norm(r.get("Nachname"))
        tel    = norm(r.get("Tel"))
        mobil  = norm(r.get("Mobil"))
        mail   = norm(r.get("E-Mail"))
        funk   = norm(r.get("Funktion/Anrede"))
        status = norm(r.get("Status"))
        firma_csv = norm(r.get("Firma CSV"))
        firma_crm = norm(r.get("Firma CRM"))
        mfid   = norm(r.get("M-Files-ID"))

        if not crm_id:
            skip_kontakte.append({
                "Grund":     "Keine CRM-ID Firma (Firma fehlt im CRM)",
                "Firma CSV": firma_csv,
                "Vorname":   vor, "Nachname": nach,
                "Tel": tel, "Mobil": mobil, "E-Mail": mail,
                "M-Files-ID": mfid,
            })
            continue

        c = by_id.get(crm_id)
        if not c:
            skip_kontakte.append({
                "Grund":     "CRM-ID nicht (mehr) im Snapshot",
                "Firma CSV": firma_csv, "Firma CRM": firma_crm,
                "Vorname":   vor, "Nachname": nach,
                "Tel": tel, "Mobil": mobil, "E-Mail": mail,
                "CRM-ID Firma": crm_id, "M-Files-ID": mfid,
            })
            continue

        if not (vor or nach):
            skip_kontakte.append({
                "Grund":     "Kein Name",
                "Firma CSV": firma_csv,
                "M-Files-ID": mfid,
            })
            continue

        if not (tel or mobil or mail):
            skip_kontakte.append({
                "Grund":     "Weder Tel noch Mail",
                "Firma CRM": c.get("company_name") or "",
                "Vorname":   vor, "Nachname": nach,
                "CRM-ID Firma": crm_id, "M-Files-ID": mfid,
            })
            continue

        existing = c.get("contact_persons") or []
        if not isinstance(existing, list): existing = []

        # Neue Kontaktperson – Schema wie es im CRM bereits verwendet wird
        new_cp = {
            "vorname": vor,
            "name":    nach,
            "phone":   tel,
            "phone2":  mobil,
            "email":   mail,
            "role":    funk,
        }

        # Kurzform für Vorschau
        existing_summary = "; ".join([
            f"{(p.get('vorname') or '').strip()} {(p.get('name') or '').strip()}".strip()
            for p in existing if isinstance(p, dict)
        ]) or "(leer)"

        update_kontakte.append({
            "CRM-ID Firma":      crm_id,
            "Firma CRM":         c.get("company_name") or "",
            "Bestehende Kontakte (vorher)": existing_summary,
            "NEU Vorname":       vor,
            "NEU Nachname":      nach,
            "NEU Tel":           tel,
            "NEU Mobil":         mobil,
            "NEU E-Mail":        mail,
            "NEU Funktion":      funk,
            "Anzahl vorher":     len(existing),
            "Anzahl nachher":    len(existing) + 1,
            "Status (Excel)":    status,
            "M-Files-ID":        mfid,
            "_JSON_Append":      json.dumps(new_cp, ensure_ascii=False),
            "_Aktion":           "UPDATE customers SET contact_persons = contact_persons || $1 WHERE id = $2",
        })

    # ── Excel schreiben ─────────────────────────────────
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    HEADER_FILL = PatternFill("solid", fgColor="5b21b6")
    HEADER_FONT = Font(color="FFFFFF", bold=True)

    def write_sheet(name, rows, cols):
        ws = wb.create_sheet(name)
        ws.append(cols)
        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(vertical="center")
        for r in rows:
            ws.append([r.get(c, "") for c in cols])
        for i, col in enumerate(cols, start=1):
            sample = [str(col)] + [str(r.get(col, "")) for r in rows[:200]]
            max_len = max(len(s) for s in sample)
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = min(max_len + 2, 60)
        ws.freeze_panes = "A2"

    # Uebersicht zuerst
    ws = wb.create_sheet("Uebersicht")
    ws.append(["Aktion", "Anzahl", "Beschreibung"])
    for cell in ws[1]:
        cell.fill = HEADER_FILL; cell.font = HEADER_FONT
    ws.append(["NEU_Privatkunden",  len(neu_privat),     "INSERT in customers (kunde_typ=privat)"])
    ws.append(["UPDATE_Kontakte",   len(update_kontakte),"contact_persons Array um neuen Kontakt erweitern"])
    ws.append(["SKIP_Kontakte",     len(skip_kontakte),  "Werden NICHT importiert (Grund siehe Sheet)"])
    ws.append([])
    ws.append(["Quelle", str(SRC_XLSX)])
    ws.append(["Snapshot", str(SNAPSHOT)])
    ws.append(["Hinweis", "Vorschau – DB wurde NICHT verändert."])
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 60

    write_sheet("NEU_Privatkunden", neu_privat, [
        "company_name", "vorname", "name", "kunde_typ",
        "phone", "phone2", "email",
        "strasse", "plz", "ort", "geburtstag",
        "M-Files-ID", "_Aktion",
    ])

    write_sheet("UPDATE_Kontakte", update_kontakte, [
        "CRM-ID Firma", "Firma CRM",
        "Bestehende Kontakte (vorher)",
        "NEU Vorname", "NEU Nachname",
        "NEU Tel", "NEU Mobil", "NEU E-Mail", "NEU Funktion",
        "Anzahl vorher", "Anzahl nachher",
        "Status (Excel)", "M-Files-ID",
        "_JSON_Append", "_Aktion",
    ])

    write_sheet("SKIP_Kontakte", skip_kontakte, [
        "Grund", "Firma CSV", "Firma CRM",
        "Vorname", "Nachname",
        "Tel", "Mobil", "E-Mail",
        "CRM-ID Firma", "M-Files-ID",
    ])

    wb.save(OUT_XLSX)
    print("", flush=True)
    print(f"[OK] Vorschau geschrieben: {OUT_XLSX}", flush=True)
    print(f"  NEU_Privatkunden:  {len(neu_privat)}", flush=True)
    print(f"  UPDATE_Kontakte:   {len(update_kontakte)}", flush=True)
    print(f"  SKIP_Kontakte:     {len(skip_kontakte)}", flush=True)
    print("", flush=True)
    print("Nichts wurde in der DB verändert. Bitte Excel prüfen.", flush=True)

if __name__ == "__main__":
    main()
