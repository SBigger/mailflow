"""
Liest Import_Vorschau.xlsx und schickt Payload an Edge Function `import-crm-data`.

Modi:
  python import_runner.py dry      → ruft Function mit dry_run=true (kein DB-Write)
  python import_runner.py go       → echter Import
  python import_runner.py          → dry (Default)

Schreibt Ergebnis nach Desktop/Import_Ergebnis.xlsx.
"""

import json
import os
import sys
from pathlib import Path

import urllib.request
import urllib.error

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

DESKTOP    = Path(r"C:\Users\SaschaBigger\Artis Treuhand GmbH\OneDrive - Artis Treuhand GmbH\Desktop")
SRC_XLSX   = DESKTOP / "Import_Vorschau.xlsx"
OUT_XLSX   = DESKTOP / "Import_Ergebnis.xlsx"

# .env aus mailflow/ – enthält SUPABASE_URL und ANON_KEY
ENV_FILES = [
    Path(__file__).resolve().parents[1] / ".env.local",
    Path(__file__).resolve().parents[1] / ".env",
]

PROJECT_REF = "uawgpxcihixqxqxxbjak"
FUNCTION_URL = f"https://{PROJECT_REF}.supabase.co/functions/v1/import-crm-data"

def load_anon_key():
    for env_file in ENV_FILES:
        if not env_file.exists(): continue
        for line in env_file.read_text(encoding="utf-8", errors="ignore").splitlines():
            line = line.strip()
            if line.startswith("VITE_SUPABASE_ANON_KEY=") or line.startswith("SUPABASE_ANON_KEY="):
                return line.split("=", 1)[1].strip().strip('"').strip("'")
    for k in ("VITE_SUPABASE_ANON_KEY", "SUPABASE_ANON_KEY"):
        v = os.environ.get(k)
        if v: return v
    raise SystemExit("Kein Supabase ANON_KEY gefunden (in .env / .env.local / ENV)")

def read_sheet_dicts(wb, name):
    if name not in wb.sheetnames: return []
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows: return []
    headers = list(rows[0])
    out = []
    for r in rows[1:]:
        if all(v in (None, "") for v in r): continue
        out.append({h: v for h, v in zip(headers, r)})
    return out

def s(v):
    if v is None: return ""
    return str(v).strip()

def build_payload():
    wb = openpyxl.load_workbook(SRC_XLSX, data_only=True)
    neu = read_sheet_dicts(wb, "NEU_Privatkunden")
    upd = read_sheet_dicts(wb, "UPDATE_Kontakte")

    inserts = []
    for r in neu:
        inserts.append({
            "company_name": s(r.get("company_name")),
            "vorname":      s(r.get("vorname")),
            "name":         s(r.get("name")),
            "kunde_typ":    s(r.get("kunde_typ")) or "privat",
            "phone":        s(r.get("phone")),
            "phone2":       s(r.get("phone2")),
            "email":        s(r.get("email")),
            "strasse":      s(r.get("strasse")),
            "plz":          s(r.get("plz")),
            "ort":          s(r.get("ort")),
            "geburtstag":   s(r.get("geburtstag")),
            "mfiles_id":    s(r.get("M-Files-ID")),
        })

    updates = []
    for r in upd:
        updates.append({
            "id": s(r.get("CRM-ID Firma")),
            "append_contact": {
                "vorname": s(r.get("NEU Vorname")),
                "name":    s(r.get("NEU Nachname")),
                "phone":   s(r.get("NEU Tel")),
                "phone2":  s(r.get("NEU Mobil")),
                "email":   s(r.get("NEU E-Mail")),
                "role":    s(r.get("NEU Funktion")),
            },
            "mfiles_id": s(r.get("M-Files-ID")),
        })

    return inserts, updates

def call_function(inserts, updates, dry_run):
    body = json.dumps({"inserts": inserts, "updates": updates, "dry_run": dry_run}).encode("utf-8")
    anon = load_anon_key()
    req = urllib.request.Request(
        FUNCTION_URL,
        data=body,
        method="POST",
        headers={
            "Content-Type": "application/json",
            "Authorization": f"Bearer {anon}",
            "apikey": anon,
        },
    )
    try:
        with urllib.request.urlopen(req, timeout=300) as resp:
            return json.loads(resp.read().decode("utf-8"))
    except urllib.error.HTTPError as e:
        return {"error": f"HTTP {e.code}", "body": e.read().decode("utf-8", errors="replace")}

def write_result(result):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    HF = PatternFill("solid", fgColor="5b21b6")
    FF = Font(color="FFFFFF", bold=True)
    OK_FILL = PatternFill("solid", fgColor="d4f5d4")
    SK_FILL = PatternFill("solid", fgColor="fff5d4")
    ER_FILL = PatternFill("solid", fgColor="ffd4d4")

    def add_sheet(name, items, cols):
        ws = wb.create_sheet(name)
        ws.append(cols)
        for c in ws[1]:
            c.fill = HF; c.font = FF; c.alignment = Alignment(vertical="center")
        for it in items:
            ws.append([it.get(c, "") for c in cols])
            row = ws[ws.max_row]
            if it.get("ok"):
                fill = OK_FILL
            elif it.get("skipped"):
                fill = SK_FILL
            else:
                fill = ER_FILL
            for c in row:
                c.fill = fill
        for i, c in enumerate(cols, 1):
            sample = [str(c)] + [str(it.get(c, "")) for it in items[:200]]
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = min(max(len(s) for s in sample) + 2, 60)
        ws.freeze_panes = "A2"

    inserts = result.get("inserts", [])
    updates = result.get("updates", [])
    summary = result.get("summary", {})

    ws = wb.create_sheet("Uebersicht")
    ws.append(["Kennzahl", "Wert"])
    for c in ws[1]: c.fill = HF; c.font = FF
    ws.append(["dry_run", str(result.get("dry_run", False))])
    for k, v in summary.items():
        ws.append([k, v])
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 14

    add_sheet("INSERTS", inserts, [
        "ok", "skipped", "action", "reason", "error",
        "id", "existing_id", "company_name", "mfiles_id",
    ])
    add_sheet("UPDATES", updates, [
        "ok", "skipped", "action", "reason", "error",
        "id", "company", "anzahl_vorher", "anzahl_nachher", "mfiles_id",
    ])

    wb.save(OUT_XLSX)
    print(f"[OK] Ergebnis: {OUT_XLSX}", flush=True)

def main():
    mode = (sys.argv[1] if len(sys.argv) > 1 else "dry").lower()
    if mode not in ("dry", "go"):
        raise SystemExit("Usage: import_runner.py [dry|go]")
    dry = (mode == "dry")

    print(f"Modus: {'DRY-RUN (kein DB-Write)' if dry else 'ECHT (schreibt in DB!)'}", flush=True)
    inserts, updates = build_payload()
    print(f"  Inserts: {len(inserts)}, Updates: {len(updates)}", flush=True)

    print("Rufe Edge Function ...", flush=True)
    result = call_function(inserts, updates, dry_run=dry)

    if "error" in result and "summary" not in result:
        print("FEHLER:", json.dumps(result, ensure_ascii=False, indent=2), flush=True)
        return

    print("Summary:", json.dumps(result.get("summary"), ensure_ascii=False), flush=True)
    write_result(result)

if __name__ == "__main__":
    main()
