"""
Abgleich M-Files-Adress-Export gegen CRM (Supabase customers).

Liest:
- C:/Users/SaschaBigger/Artis Treuhand GmbH/OneDrive - Artis Treuhand GmbH/Desktop/alle adressen m-files.CSV
- C:/tmp/crm_snapshot.json (vom export-crm-snapshot Edge Function)

Schreibt:
- C:/Users/SaschaBigger/Artis Treuhand GmbH/OneDrive - Artis Treuhand GmbH/Desktop/Abgleich_Adressen.xlsx

Sheets:
  1. Firmenkunden_ergaenzen       (im CRM vorhanden, CSV liefert mehr Tel/Mail)
  2. Firmenkunden_fehlen          (im CSV, im CRM nicht gefunden)
  3. Privatkunden_ergaenzen
  4. Privatkunden_fehlen
  5. Kontakte_ergaenzen           (Personen → Firma im CRM, fehlen oder unvollständig)

Es wird NICHTS geschrieben oder importiert. Reines Read+Excel.
"""

import csv
import json
import re
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

# ── Pfade ───────────────────────────────────────────────
CSV_PATH      = Path(r"C:\Users\SaschaBigger\Artis Treuhand GmbH\OneDrive - Artis Treuhand GmbH\Desktop\alle adressen m-files.CSV")
import os
SNAPSHOT_PATH = Path(os.environ.get("TEMP", r"C:\Users\SaschaBigger\AppData\Local\Temp")) / "crm_snapshot.json"

# ── Konfiguration: was generell NICHT importiert werden soll ─────────
# Substring-Match (case-insensitive) gegen Firmenname / Vor+Nachname
EXCLUDE_PATTERNS = [
    "revisio",            # revisio treuhand gmbh – nicht übernehmen
    "interrevision",      # Audit-Firma
    # Hier kommen weitere dazu, die du explizit ausschliessen willst
]

def is_excluded(text):
    t = norm_lower(text)
    return any(p in t for p in EXCLUDE_PATTERNS)
OUT_PATH      = Path(r"C:\Users\SaschaBigger\Artis Treuhand GmbH\OneDrive - Artis Treuhand GmbH\Desktop\Abgleich_Adressen.xlsx")

# ── Helpers ─────────────────────────────────────────────
def norm(s):
    if s is None: return ""
    s = str(s).strip()
    s = re.sub(r"\s+", " ", s)
    return s

def norm_lower(s):
    return norm(s).lower()

def norm_phone(s):
    if not s: return ""
    digits = re.sub(r"\D", "", str(s))
    if not digits: return ""
    if digits.startswith("00"): digits = digits[2:]
    if digits.startswith("41") and len(digits) >= 11:
        return "+" + digits
    if digits.startswith("0"):
        return "+41" + digits[1:]
    return "+" + digits if not digits.startswith("+") else digits

def phone_suffix(s):
    d = re.sub(r"\D", "", str(s or ""))
    return d[-9:] if len(d) >= 7 else ""

def fmt_phone(s):
    """Schweizer/Deutsche Telefonnummern in einheitliche Form bringen.

    Regeln (Konvention Sascha):
      - CH/Inland: '071 234 56 78'  (Beginn mit 0)
      - International (CH): '+41 71 234 56 78'
      - International (DE): '+49 30 12345 6789'
      - Sonst: '+CC ...' so wie geliefert.
    Leerer Input → ''.
    """
    if not s: return ""
    raw = str(s).strip().lstrip("'").strip()
    if not raw: return ""
    # Trennzeichen herausnehmen
    digits = re.sub(r"\D", "", raw)
    if not digits: return ""
    # 00-Präfix → +
    has_plus = raw.startswith("+") or digits.startswith("00")
    if digits.startswith("00"): digits = digits[2:]; has_plus = True

    # Wenn 9-10 Ziffern und beginnt mit 0 → Schweizer Inland-Format
    if not has_plus and digits.startswith("0") and 9 <= len(digits) <= 11:
        d = digits
        # Format: 0XX XXX XX XX  (10-stellig CH)
        if len(d) == 10:
            return f"{d[0:3]} {d[3:6]} {d[6:8]} {d[8:10]}"
        return d  # Sonderfall

    # +41 (Schweiz)
    if has_plus and digits.startswith("41"):
        rest = digits[2:].lstrip("0")  # +41 darf danach keine 0 haben
        if len(rest) == 9:
            return f"+41 {rest[0:2]} {rest[2:5]} {rest[5:7]} {rest[7:9]}"
        return f"+41 {rest}"

    # +49 (Deutschland)
    if has_plus and digits.startswith("49"):
        rest = digits[2:].lstrip("0")
        if len(rest) >= 7:
            # grobes Format: +49 XX XXXX XXXX
            return f"+49 {rest[0:2]} {rest[2:6]} {rest[6:]}"
        return f"+49 {rest}"

    # Andere Internationale → +CC <rest>
    if has_plus:
        return "+" + digits

    # Fallback: keine Erkennung → roh
    return raw

def norm_email(s):
    return norm_lower(s)

def norm_company(s):
    s = norm_lower(s)
    # AG / GmbH / & Co. Klammern entfernen
    s = re.sub(r"[.,]", "", s)
    s = re.sub(r"\s+&\s+", " und ", s)
    s = re.sub(r"\s+", " ", s)
    return s.strip()

def norm_person(vor, nach):
    return norm_lower(vor) + "|" + norm_lower(nach)

# ── CSV laden ──────────────────────────────────────────
def load_csv():
    rows = []
    with open(CSV_PATH, encoding="utf-8-sig") as f:
        r = csv.DictReader(f, delimiter=";")
        for row in r:
            rows.append(row)
    return rows

# ── CRM snapshot laden ─────────────────────────────────
def load_crm():
    with open(SNAPSHOT_PATH, encoding="utf-8") as f:
        data = json.load(f)
    return data.get("customers", [])

# ── Map building ───────────────────────────────────────
def build_crm_indexes(customers):
    by_company = {}        # norm_company → customer
    by_person  = {}        # vor|nach → customer
    by_phone   = {}        # phone_suffix → set(ids)
    by_email   = {}        # email_lower → set(ids)
    for c in customers:
        comp = norm_company(c.get("company_name") or "")
        if comp:
            by_company.setdefault(comp, []).append(c)
        person_key = norm_person(c.get("vorname"), c.get("name"))
        if person_key.strip("|"):
            by_person.setdefault(person_key, []).append(c)
        ps = phone_suffix(c.get("phone"))
        if ps:
            by_phone.setdefault(ps, set()).add(c["id"])
        em = norm_email(c.get("email"))
        if em:
            by_email.setdefault(em, set()).add(c["id"])
    return by_company, by_person, by_phone, by_email

# ── CSV-Zeilen aufbereiten ─────────────────────────────
def csv_phones(row):
    """Liefert (best_phone, mobile, alle_phones_string) – formatiert."""
    z = fmt_phone(row.get("Telefon (Zentrale)"))
    g = fmt_phone(row.get("Tel 1 (Geschäft)"))
    p = fmt_phone(row.get("Tel 2 (Privat)"))
    m = fmt_phone(row.get("Tel 3 (Mobil)"))
    # Primary: Direktnummer Geschäft bevorzugt vor Zentrale
    primary = g or z or p or m
    parts = []
    if z: parts.append(f"Zentrale: {z}")
    if g: parts.append(f"Direkt:   {g}")
    if p: parts.append(f"Privat:   {p}")
    if m: parts.append(f"Mobil:    {m}")
    return primary, m, "  |  ".join(parts)

def csv_address_full(row):
    parts = []
    str_h = " ".join([norm(row.get("Strasse")), norm(row.get("Hausnummer"))]).strip()
    if str_h: parts.append(str_h)
    plz_ort = " ".join([norm(row.get("PLZ")), norm(row.get("Ort"))]).strip()
    if plz_ort: parts.append(plz_ort)
    return ", ".join(parts)

def crm_addr_str(c):
    return ", ".join(filter(None, [
        norm(c.get("strasse")),
        " ".join(filter(None, [norm(c.get("plz")), norm(c.get("ort"))])).strip()
    ]))

# ── Abgleich-Funktionen ────────────────────────────────
def diff_field(crm_val, csv_val, *, is_phone=False, is_email=False):
    """Was sollte ergänzt werden?
       Liefert (leer_im_crm: bool, csv_neu: bool, csv_value)
    """
    crm_n = norm(crm_val)
    csv_n = norm(csv_val)
    if not csv_n:
        return False, False, ""
    if not crm_n:
        return True, True, csv_n
    if is_phone:
        if phone_suffix(crm_n) == phone_suffix(csv_n):
            return False, False, csv_n
        return False, True, csv_n  # CRM hat anderen Wert → Vorschlag zum Vergleich
    if is_email:
        if norm_lower(crm_n) == norm_lower(csv_n):
            return False, False, csv_n
        return False, True, csv_n
    return False, False, csv_n

# ── Hauptauswertung ────────────────────────────────────
def main():
    print(f"Lade CSV ...", flush=True)
    csv_rows = load_csv()
    print(f"  {len(csv_rows)} Zeilen", flush=True)

    print(f"Lade CRM snapshot ...", flush=True)
    customers = load_crm()
    print(f"  {len(customers)} Kunden", flush=True)

    by_company, by_person, by_phone, by_email = build_crm_indexes(customers)

    firmen_csv  = [r for r in csv_rows if norm(r.get("Klasse")) == "Firmenkunde"]
    privat_csv  = [r for r in csv_rows if norm(r.get("Klasse")) == "Privatkunde"]
    kontakt_csv = [r for r in csv_rows if norm(r.get("Klasse")) == "Kontakt"]

    print(f"  CSV: {len(firmen_csv)} Firmen, {len(privat_csv)} Privat, {len(kontakt_csv)} Kontakte", flush=True)

    # ── Bereits rot markierte M-Files-IDs aus existierendem Excel lesen ──
    excluded_mfiles_ids = set()
    if OUT_PATH.exists():
        try:
            wb_old = openpyxl.load_workbook(OUT_PATH)
            for sname in wb_old.sheetnames:
                ws = wb_old[sname]
                # Header lesen
                headers = [c.value for c in ws[1]]
                if "M-Files-ID" not in headers: continue
                idx_mfid = headers.index("M-Files-ID")
                for row in ws.iter_rows(min_row=2):
                    is_red = False
                    for cell in row:
                        f = cell.fill
                        if f and f.fgColor and f.fgColor.rgb:
                            rgb = str(f.fgColor.rgb)
                            try:
                                rr = int(rgb[2:4], 16); gg = int(rgb[4:6], 16); bb = int(rgb[6:8], 16)
                                if rr > 200 and gg < 180 and bb < 180 and (rr - gg) > 40:
                                    is_red = True; break
                            except Exception:
                                pass
                    if is_red:
                        mfid = row[idx_mfid].value
                        if mfid: excluded_mfiles_ids.add(str(mfid))
            if excluded_mfiles_ids:
                print(f"  Aus existierendem Excel: {len(excluded_mfiles_ids)} rot markierte Zeilen werden übersprungen", flush=True)
        except Exception as e:
            print(f"  (alte Excel konnte nicht gelesen werden: {e})", flush=True)

    def skip_by_id(r):
        return str(norm(r.get("ID"))) in excluded_mfiles_ids

    # Firmenkunden ergänzen + fehlen
    firma_ergaenzen = []
    firma_fehlen    = []
    for r in firmen_csv:
        if skip_by_id(r): continue
        firmenname = norm(r.get("Firmenname")) or norm(r.get("Name"))
        if not firmenname: continue
        if is_excluded(firmenname): continue
        key = norm_company(firmenname)
        matches = by_company.get(key, [])
        # auch fuzzy: enthält
        if not matches:
            for k, v in by_company.items():
                if key and (key in k or k in key) and len(key) > 8:
                    matches.extend(v); break
        if not matches:
            phone, mobile, all_phones = csv_phones(r)
            firma_fehlen.append({
                "Firmenname":   firmenname,
                "Adresse":      csv_address_full(r),
                "Telefon":      phone,
                "E-Mail":       norm(r.get("E-Mail")),
                "Alle Tel":     all_phones,
                "Bemerkung":    norm(r.get("Bemerkungen / Notiz")),
                "M-Files-ID":   norm(r.get("ID")),
            })
            continue
        c = matches[0]
        phone, mobile, all_phones = csv_phones(r)
        crm_phone = fmt_phone(c.get("phone"))
        crm_email = norm(c.get("email"))
        leer_phone, vorschlag_phone, _ = diff_field(crm_phone, phone, is_phone=True)
        leer_email, vorschlag_email, _ = diff_field(crm_email, norm(r.get("E-Mail")), is_email=True)
        if leer_phone or leer_email or (vorschlag_phone and not crm_phone):
            firma_ergaenzen.append({
                "CRM-ID":         c["id"],
                "Firmenname CRM": c.get("company_name") or "",
                "Firmenname CSV": firmenname,
                "Tel CRM":        crm_phone,
                "Tel CSV (neu)":  phone if leer_phone else "",
                "Tel CSV (anders)": phone if (vorschlag_phone and crm_phone) else "",
                "Mail CRM":       crm_email,
                "Mail CSV (neu)": norm(r.get("E-Mail")) if leer_email else "",
                "Mail CSV (anders)": norm(r.get("E-Mail")) if (vorschlag_email and crm_email) else "",
                "Alle Tel CSV":   all_phones,
                "M-Files-ID":     norm(r.get("ID")),
            })

    # Privatkunden ergänzen + fehlen
    # Wichtig: im CRM sind Privatkunden meist als customer mit
    # company_name = "Vorname Nachname" angelegt (vorname gesetzt, name oft NULL).
    privat_ergaenzen = []
    privat_fehlen    = []
    for r in privat_csv:
        if skip_by_id(r): continue
        vor = norm(r.get("Vorname"))
        nach = norm(r.get("Nachname"))
        if not (vor or nach): continue
        if is_excluded(f"{vor} {nach}"): continue
        key = norm_person(vor, nach)
        matches = by_person.get(key, [])
        # Fallback 1: company_name = "Vorname Nachname"
        if not matches:
            key_a = norm_company(f"{vor} {nach}")
            matches = by_company.get(key_a, [])
        # Fallback 2: company_name = "Nachname Vorname"
        if not matches:
            key_b = norm_company(f"{nach} {vor}")
            matches = by_company.get(key_b, [])
        # Fallback 3: vorname allein und Nachname matched anders
        if not matches:
            for c in customers:
                cn = norm_lower(c.get("company_name") or "")
                if not cn: continue
                if (norm_lower(vor) in cn) and (norm_lower(nach) in cn):
                    matches = [c]; break
        if not matches:
            phone, mobile, all_phones = csv_phones(r)
            privat_fehlen.append({
                "Vorname":     vor,
                "Nachname":    nach,
                "Adresse":     csv_address_full(r),
                "Telefon":     phone,
                "Mobil":       mobile,
                "E-Mail":      norm(r.get("E-Mail")),
                "Alle Tel":    all_phones,
                "Geburtstag":  norm(r.get("Geburtstag")),
                "M-Files-ID":  norm(r.get("ID")),
            })
            continue
        c = matches[0]
        phone, mobile, all_phones = csv_phones(r)
        crm_phone = fmt_phone(c.get("phone"))
        crm_email = norm(c.get("email"))
        leer_phone, _, _ = diff_field(crm_phone, phone, is_phone=True)
        leer_email, _, _ = diff_field(crm_email, norm(r.get("E-Mail")), is_email=True)
        if leer_phone or leer_email:
            privat_ergaenzen.append({
                "CRM-ID":         c["id"],
                "Name CRM":       (c.get("vorname") or "") + " " + (c.get("name") or ""),
                "Name CSV":       vor + " " + nach,
                "Tel CRM":        crm_phone,
                "Tel CSV (neu)":  phone if leer_phone else "",
                "Mail CRM":       crm_email,
                "Mail CSV (neu)": norm(r.get("E-Mail")) if leer_email else "",
                "Alle Tel CSV":   all_phones,
                "M-Files-ID":     norm(r.get("ID")),
            })

    # ── Helper: Firma-Referenz(en) aus Adressen-Spalte ────────────
    # Spaltenformat: "Aerne AG" oder "Foo AG (gelöscht)" oder
    # "Banderet AG; Maba Holding AG; Maba Immo AG"
    def parse_firma_refs(s):
        if not s: return []
        # Mehrere durch ";" oder Komma trennen (Komma vorsichtig, nur wenn nicht " AG, ...")
        parts = re.split(r"\s*;\s*", s)
        out = []
        for p in parts:
            p = p.strip().strip("'").strip('"')
            if not p: continue
            # "(gelöscht)" markierung extrahieren
            geloescht = bool(re.search(r"\(gel.?scht\)", p, re.IGNORECASE))
            p_clean = re.sub(r"\s*\(gel.?scht\)\s*$", "", p, flags=re.IGNORECASE).strip()
            if not p_clean: continue
            # Generische Sammel-Adresse ignorieren
            if p_clean.lower() in ("laufkunden", "laufkunde", "diverse"):
                continue
            out.append({"name": p_clean, "geloescht": geloescht})
        return out

    def find_firma_in_crm(firma_name):
        key = norm_company(firma_name)
        matches = by_company.get(key, [])
        if matches: return matches[0]
        # Fuzzy: enthält / ist enthalten
        if key and len(key) > 5:
            for k, v in by_company.items():
                if (key in k) or (k in key):
                    return v[0]
        # Token-basiert: wenn alle Tokens > 4 Zeichen vorkommen
        tokens = [t for t in re.split(r"\s+", key) if len(t) > 4 and t not in ("treuhand","holding","invest")]
        if tokens:
            for k, v in by_company.items():
                if all(t in k for t in tokens):
                    return v[0]
        return None

    # Kontakte: über "Adressen"-Spalte → Firmenname(n) referenziert
    kontakte_ergaenzen = []
    kontakte_ohne_firma = []
    for r in kontakt_csv:
        if skip_by_id(r): continue
        vor = norm(r.get("Vorname"))
        nach = norm(r.get("Nachname"))
        if is_excluded(f"{vor} {nach}"): continue
        if is_excluded(norm(r.get("Adressen"))): continue
        # Primärquelle: "Adressen" (Plural!) – das ist die M-Files Verbindung
        firmen_refs = parse_firma_refs(norm(r.get("Adressen")))
        # Fallback nur wenn "Adressen" leer
        if not firmen_refs:
            firmen_refs = parse_firma_refs(norm(r.get("Adresse")))

        phone, mobile, all_phones = csv_phones(r)
        email = norm(r.get("E-Mail"))
        funktion = norm(r.get("Anrede")) or norm(r.get("Anrede Briefkopf"))

        if not firmen_refs:
            kontakte_ohne_firma.append({
                "Vorname": vor, "Nachname": nach,
                "Telefon": phone, "Mobil": mobile, "E-Mail": email,
                "M-Files-ID": norm(r.get("ID")),
            })
            continue

        # Erste nicht-gelöschte Firma bevorzugen
        primary = None
        for ref in firmen_refs:
            if not ref["geloescht"]:
                primary = ref; break
        if primary is None:
            primary = firmen_refs[0]  # alle gelöscht → trotzdem Versuch
        firmenname_csv = primary["name"]
        is_geloescht = primary["geloescht"]

        c = find_firma_in_crm(firmenname_csv)

        if not c:
            kontakte_ergaenzen.append({
                "Status":          "Firma gelöscht (M-Files)" if is_geloescht else "Firma fehlt im CRM",
                "Firma CSV":       firmenname_csv + (" [gelöscht]" if is_geloescht else ""),
                "Firma CRM":       "",
                "Vorname":         vor,
                "Nachname":        nach,
                "Funktion/Anrede": funktion,
                "Tel":             phone,
                "Mobil":           mobile,
                "E-Mail":          email,
                "Alle Tel":        all_phones,
                "Im CRM bereits?": "—",
                "CRM-ID Firma":    "",
                "M-Files-ID":      norm(r.get("ID")),
            })
            continue

        existing = c.get("contact_persons") or []
        if not isinstance(existing, list): existing = []

        # Schon vorhanden? (Vor+Nachname matcht)
        already = None
        for cp in existing:
            if norm_lower(cp.get("vorname")) == norm_lower(vor) and norm_lower(cp.get("name")) == norm_lower(nach):
                already = cp; break
        if already is None:
            kontakte_ergaenzen.append({
                "Status":          "fehlt als Kontaktperson",
                "Firma CSV":       firmenname_csv,
                "Firma CRM":       c.get("company_name") or "",
                "Vorname":         vor,
                "Nachname":        nach,
                "Funktion/Anrede": funktion,
                "Tel":             phone,
                "Mobil":           mobile,
                "E-Mail":          email,
                "Alle Tel":        all_phones,
                "Im CRM bereits?": "Nein",
                "CRM-ID Firma":    c["id"],
                "M-Files-ID":      norm(r.get("ID")),
            })
        else:
            ex_phone = fmt_phone(already.get("phone"))
            ex_phone2 = fmt_phone(already.get("phone2"))
            ex_email = norm(already.get("email"))
            leer_phone = bool(phone) and not (ex_phone or ex_phone2)
            leer_email = bool(email) and not ex_email
            if leer_phone or leer_email:
                kontakte_ergaenzen.append({
                    "Status":          "Kontakt ohne Tel/Mail",
                    "Firma CSV":       firmenname_csv,
                    "Firma CRM":       c.get("company_name") or "",
                    "Vorname":         vor,
                    "Nachname":        nach,
                    "Funktion/Anrede": funktion,
                    "Tel":             phone if leer_phone else "",
                    "Mobil":           mobile if leer_phone else "",
                    "E-Mail":          email if leer_email else "",
                    "Alle Tel":        all_phones,
                    "Im CRM bereits?": f"Ja (Tel:{ex_phone or '-'} / Mail:{ex_email or '-'})",
                    "CRM-ID Firma":    c["id"],
                    "M-Files-ID":      norm(r.get("ID")),
                })

    # ── Excel schreiben ────────────────────────────────
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    HEADER_FILL = PatternFill("solid", fgColor="5b21b6")
    HEADER_FONT = Font(color="FFFFFF", bold=True)

    def write_sheet(name, rows, cols):
        ws = wb.create_sheet(name)
        ws.append(cols)
        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(vertical="center")
        for r in rows:
            ws.append([r.get(c, "") for c in cols])
        # Spaltenbreiten
        for i, col in enumerate(cols, start=1):
            max_len = max([len(str(col))] + [len(str(r.get(col, ""))) for r in rows[:200]])
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = min(max_len + 2, 50)
        ws.freeze_panes = "A2"

    write_sheet("Firmenkunden_ergaenzen", firma_ergaenzen, [
        "Firmenname CRM", "Firmenname CSV",
        "Tel CRM", "Tel CSV (neu)", "Tel CSV (anders)",
        "Mail CRM", "Mail CSV (neu)", "Mail CSV (anders)",
        "Alle Tel CSV", "M-Files-ID", "CRM-ID",
    ])
    write_sheet("Firmenkunden_fehlen", firma_fehlen, [
        "Firmenname", "Adresse", "Telefon", "E-Mail", "Alle Tel", "Bemerkung", "M-Files-ID",
    ])
    write_sheet("Privatkunden_ergaenzen", privat_ergaenzen, [
        "Name CRM", "Name CSV", "Tel CRM", "Tel CSV (neu)", "Mail CRM", "Mail CSV (neu)",
        "Alle Tel CSV", "M-Files-ID", "CRM-ID",
    ])
    write_sheet("Privatkunden_fehlen", privat_fehlen, [
        "Vorname", "Nachname", "Adresse", "Telefon", "Mobil", "E-Mail", "Alle Tel",
        "Geburtstag", "M-Files-ID",
    ])
    write_sheet("Kontakte_ergaenzen", kontakte_ergaenzen, [
        "Status", "Firma CSV", "Firma CRM", "Vorname", "Nachname", "Funktion/Anrede",
        "Tel", "Mobil", "E-Mail", "Alle Tel", "Im CRM bereits?", "CRM-ID Firma", "M-Files-ID",
    ])
    if kontakte_ohne_firma:
        write_sheet("Kontakte_ohne_Firmenref", kontakte_ohne_firma,
                    ["Vorname", "Nachname", "Telefon", "Mobil", "E-Mail", "M-Files-ID"])

    # Übersichts-Sheet
    ws = wb.create_sheet("Uebersicht", 0)
    ws.append(["Sheet", "Zeilen"])
    for cell in ws[1]:
        cell.fill = HEADER_FILL; cell.font = HEADER_FONT
    ws.append(["Firmenkunden_ergaenzen", len(firma_ergaenzen)])
    ws.append(["Firmenkunden_fehlen", len(firma_fehlen)])
    ws.append(["Privatkunden_ergaenzen", len(privat_ergaenzen)])
    ws.append(["Privatkunden_fehlen", len(privat_fehlen)])
    ws.append(["Kontakte_ergaenzen", len(kontakte_ergaenzen)])
    ws.append(["Kontakte_ohne_Firmenref", len(kontakte_ohne_firma)])
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 12

    wb.save(OUT_PATH)
    print(f"\n[OK] Excel: {OUT_PATH}", flush=True)
    print(f"  Firmen ergänzen:   {len(firma_ergaenzen)}", flush=True)
    print(f"  Firmen fehlen:     {len(firma_fehlen)}", flush=True)
    print(f"  Privat ergänzen:   {len(privat_ergaenzen)}", flush=True)
    print(f"  Privat fehlen:     {len(privat_fehlen)}", flush=True)
    print(f"  Kontakte erg.:     {len(kontakte_ergaenzen)}", flush=True)
    print(f"  Kontakte ohne Fa:  {len(kontakte_ohne_firma)}", flush=True)

if __name__ == "__main__":
    main()
